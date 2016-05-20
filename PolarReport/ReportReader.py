#
# Monthly report parsing
#
import re
import datetime
import sys, errno
from pprint import pprint

try:
    import openpyxl
except:
    exit("Virhe: openpyxl moduulia ei loydy")

COVERSHEETNAME = "Yleistiedot"
DATASHEETNAME = "Kuukausi-ilmoitus"
COVER_DATA_COL = 4
COVER_DATE_ROW = 5
COVER_NAME_ROW = 7
COVER_EMAIL_ROW = 8
PROJECT_NUMBER_COL = 2
ABSENCE_COL = 2
WORK_DESC_COL = 4
ORDER_NUMBER_COL = 6
DATE_ROW = 7

def readReport(fileName):
    return extractData(openWorkbook(fileName))

def openWorkbook(fileName):
    wb = None

    try:
        print "Avataan Excel-tiedosto: " + fileName
        wb = openpyxl.load_workbook(fileName, data_only=True)
    except:
        print "Virhe: Tiedoston avaaminen epaonnistui: " + fileName

    return wb

def extractData(wb):
    if not wb:
        return

    data = {}

    try:
        sheet = wb.get_sheet_by_name(COVERSHEETNAME)
    except:
        print "Virhe: Yleistiedot-valilehtea ei loydy"
        return
    
    data['raportointipaiva'] = sheet.cell(row=COVER_DATE_ROW, column=COVER_DATA_COL).value
    data['henkilo'] = sheet.cell(row=COVER_NAME_ROW, column=COVER_DATA_COL).value
    data['sposti'] = sheet.cell(row=COVER_EMAIL_ROW, column=COVER_DATA_COL).value

    if not (data['henkilo'] or data['sposti']):
        print "Virhe: henkilotiedot puuttuu Yleistiedot-valilehdelta"
        return

    if "sukunimi" in data['sposti']:
        print "Virhe: sposti virheellinen!"

    sheet = wb.get_sheet_by_name(DATASHEETNAME)
    if not sheet:
        print "Virhe: tunti-ilmoitus valilehtea ei loydy"
        return

    project_rows = indexProjects(sheet)

    absence_rows = indexAbsences(sheet)

    date_cols = indexDates(sheet)

    data['entries'] = workEntries(sheet, project_rows, date_cols)
        
    data['absenceEntries'] = absenceEntries(sheet, absence_rows, date_cols)

    return data

def indexProjects(sheet):
    rows = []
    
    for i in range(1,300):
        txt = sheet.cell(row=i, column=PROJECT_NUMBER_COL).value
        
        if txt:
            if re.match('^\d{4}$', unicode(txt)):
                rows.append(i)

    return rows

def indexAbsences(sheet):
    rows = []
    
    for i in range(1,300):
        txt = sheet.cell(row=i, column=ABSENCE_COL).value
        
        if txt:
            if re.match('Muu poissaolo', unicode(txt)):
                rows.append(i)

            if re.match('Sairasloma', unicode(txt)):
                rows.append(i)
               
            if re.match('Loma', unicode(txt)):
                rows.append(i)

    return rows

def indexDates(sheet):
    cols = []

    for i in range(1,50):
        dt = sheet.cell(row=DATE_ROW, column=i).value
        
        if dt and isinstance(dt, datetime.datetime):
            cols.append(i)

    return cols

def workEntries(sheet, rows, cols):
    entries = []

    for i in rows:
        project = int(sheet.cell(row=i, column=PROJECT_NUMBER_COL).value)
        description = sheet.cell(row=i, column=WORK_DESC_COL).value
        order = sheet.cell(row=i, column=ORDER_NUMBER_COL).value

        for j in cols:
            entry = {}
            
            fields = ['norm', 'lisatyo', 'ylit50', 'ylit100', 'ylit150',\
                      'ylit200', 'viikkoylit', 'matkat', 'km', 'ateria',\
                      'osapaivar', 'paivaraha']

            for k in range(0, len(fields)):
                v = sheet.cell(row=i+k, column=j).value
                if v:
                    entry[fields[k]] = v

            if entry:
                entry['tyonumero'] = project
                entry['tilaus'] = order if order else ''
                entry['suorituspaiva'] = sheet.cell(row=DATE_ROW, column=j).value
                entry['tyoselite'] = description if description else ''
                
                entries.append(entry)

    return entries

def absenceEntries(sheet, rows, cols):
    entries = []

    for i in rows:
        project = '0000'
        description = sheet.cell(row=i, column=ABSENCE_COL).value
        
        for j in cols:
            entry = {}
            
            v = sheet.cell(row=i, column=j).value
            if v:
                entry['norm'] = v

            if entry:
                entry['tyonumero'] = project
                entry['tilaus'] = ''
                entry['suorituspaiva'] = sheet.cell(row=DATE_ROW, column=j).value
                entry['tyoselite'] = description if description else ''
                
                entries.append(entry)

    return entries


# TEST
if __name__ == "__main__":
    file = ".\\TestData\\2016\\March\\report_1.xlsx"

    wb = openWorkbook(file)
    
    data = extractData(wb)
    pprint(data)
