#
# Generates attachments for invoicing
#
import re
import datetime
import sys, errno
from pprint import pprint

try:
    import openpyxl
except:
    exit("ERROR: Importing openpyxl failed")

COVERSHEETNAME = "Yleistiedot"
DATASHEETNAME = "Kuukausi-ilmoitus"
COVER_DATA_COL = 4
COVER_NAME_ROW = 7
COVER_EMAIL_ROW = 8
PROJECT_NUMBER_COL = 2
WORK_DESC_COL = 4
ORDER_NUMBER_COL = 6
DATE_ROW = 7

def openExcelWorkbook(fileName):
    wb = None

    try:
        print "Open excel workbook: " + fileName
        wb = openpyxl.load_workbook(fileName, data_only=True)
    except:
        exit("Virhe: Excel-tiedoston avaaminen epaonnistui: " + fileName)
    else:
        return wb

def getDates(wb):
    if not wb:
        return

    sheet = wb.get_sheet_by_name(DATASHEETNAME)
    if not sheet:
        print "Virhe: yleistiedot valilehtea ei loydy"

    return indexDates(sheet)

def indexDates(sheet):
    dates = {}

    for i in range(1,50):
        dt = sheet.cell(row=DATE_ROW, column=i).value
        print dt
        if dt and isinstance(dt, datetime.datetime):
            dates[i] = dt

    return dates


# TEST
if __name__ == "__main__":
    file = ".\\TestData\\Invoicing\\template.xlsx"

    wb = openExcelWorkbook(file)
    
    data = getDates(wb)
    pprint(data)